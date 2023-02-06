'''
methanol_RCC.py

Generates a TEA for flue gas CO2 capture from an NGCC power plant and conversion
to methanol using hydrogen generated from a hybrid wind/solar plant nearby.
Uses NREL Annual Technology Baseline (ATB) future tech development scenarios.
Tracks all of the material streams and exports in a convenient format for LCA.

Author: Jonathan Martin <jonathan.martin@nrel.gov>
Uses ASPEN process model developed by Eric Tan <eric.tan@nrel.gov>

'''

## Import libraries needed
#region

import json
import numpy as np
from pathlib import Path
from global_land_mask import globe
import matplotlib.pyplot as plt
from scipy.optimize import curve_fit

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

from hybrid.sites import SiteInfo
from hybrid.hybrid_simulation import HybridSimulation
from hybrid.log import hybrid_logger as logger
from hybrid.keys import set_nrel_key_dot_env

from examples.H2_Analysis.H2AModel_capex_opex_scenarios import H2AModel_costs

#endregion

## Convert dollars to correct year
#region

def inflate(dollars, original_year, new_year):

    # TODO: Adjust dollars using CPI
    standard_inflation = 0.025
    dollars += (new_year-original_year)**(1+standard_inflation)

    return dollars
    
#endregion    

## ALL CHANGES HERE - SELECT SCENARIO COMBOS, MANUALLY OVERRIDE PRICING ASSUMPTIONS
#region

sim_basis_year = 2020
sim_start_year = 2020
sim_end_year = 2050
sim_years = np.arange(sim_start_year,sim_end_year)

plant_lifespan = 30 # years
discount_rate = 0.07 # (= fixed charge rate)
TASC_multiplier = 1.093 # total overnight cost * TASC multiplier = total as-spent cost

NG_price_MMBTU = 4.56 # 2020 $/MMBTU TODO: make variable NG price scenarios
H2O_price_Tgal = 2.56 # 2020 $/Tgal TODO: make variable water price scenarios

select_locations = False # Switch to True to only analyze locations listed below
locations = {'IA': {'on_land':[True ,], 'lat':[43.094000,], 'lon':[-93.292220,]},
             'TX': {'on_land':[True ,], 'lat':[32.337679,], 'lon':[-97.734610,]},
             'NJ': {'on_land':[False ,], 'lat':[39.600000,], 'lon':[-73.400000,]},
             }

manual_LCOE = False # Switch to True to override HOPP-calculated LCOE
LCOE_kWh = 0.04 # $/kWh
manual_LCOH = False # Switch to True to override HOPP-calculated LCOE
LCOH_kg = 0.02 # $/kg

# Switch dict definitions to use certain scenarios for different technologies
# Options for non-H2, non-MeOH: 'Conservative', 'Moderate', 'Advanced' (from ATB)
# Options for H2: 'Current', 'Future': 'Current' holds 2015 baseline H2A scenario,
#       'Future' interpolates between that and 2040 'Future' H2A scenario
# Options for MeOH: 'Baseline CO2', others to come as ASPEN models developed
atb_scenarios = ['Advanced','Moderate','Conservative']
H2A_scenarios = ['Current','Future']
MeOH_scenarios = ['Baseline CO2 Capture'] # TBA: 'NREL CO2 Capture', 'NREL Flue Gas'
plant_scenarios =  {'NGCC':'Conservative',
                    'CCS': 'Conservative',
                    'PV':  'Advanced',
                    'LBW': 'Advanced',
                    'OSW': 'Advanced',
                    'H2':  'Future',
                    'MeOH':'Baseline CO2 Capture'}
scenario_info = {'plant_scenarios': plant_scenarios,
                'atb_scenarios':    atb_scenarios,
                'H2A_scenarios':    H2A_scenarios,
                'MeOH_scenarios':   MeOH_scenarios} 

min_plant_dist = 120 # km, minimum distance between NGCC plants in survey
land_rad = 60 # km radius of survey area around NGCC plant on land
osw_rad = 20 # km radium of survey area around offshore wind location

NGCC_cap = 100 # MW, output of NGCC plant to scale results to
NGCC_Cf = 0.85 # capacity factor of NGCC plant to scale results to
''' NGCC_cap*NGCC_Cf must be <170 MW for H2A model scaling to stay valid!
    (H2 output needs to stay below 200,000 kg H2/day'''

H2_Cf = 0.97 # capacity factor of H2 plants to scale results to

resource_dir = Path(__file__).parent.absolute()/'resource_files'/'methanol_RCC'

#endregion

## Physical constants
#region

C_MW = 12.011 # g/mol C
H_MW = 1.008 # g/mol H
O_MW = 15.999 # g/mol O
L_gal = 3.78541 # L/gal
kg_lb = 0.453592 # kg/lb
kJ_BTU = 1.05506 # kJ/BTU

MeOH_LHV_MJ_kg = 20.1 # Methanol net calorific value, MJ/kg
NG_LHV_MJ_kg = 47.1 # Natural gas net calorific value, MJ/kg
H2_LHV_MJ_kg = 120 # Hygrogen net calorific value, MJ/kg

#endregion

## Import the NREL ATB scenario pricing info
# TODO: Break out BOS costs (not in spreadsheets...)
#region

# Set constants not given in spreadsheet
basis_year = 2020
pv_constants = {'array_type':2,
                'azimuth':180,
                'inv_eff':None,
                'dc_ac_ratio':1.28} # ILR NOT OPTIMIZED - can HOPP do this?
bos_ratio = {'LBW': 322/1462, 
             'PV': 20/89}
wind_constants =   {'wind_turbine_max_cp':0.55,
                    'avail_bop_loss':0,
                    'avail_grid_loss':0,
                    'avail_turb_loss':0,
                    'elec_eff_loss':0,
                    'elec_parasitic_loss':0,
                    'env_degrad_loss':0,
                    'env_env_loss':0,
                    'env_icing_loss':0,
                    'ops_env_loss':0,
                    'ops_grid_loss':0,
                    'ops_load_loss':0,
                    'turb_generic_loss':0,
                    'turb_hysteresis_loss':0,
                    'turb_perf_loss':0,
                    'turb_specific_loss':0,
                    'wake_ext_loss':0}
osw_constants =   {'wind_turbine_max_cp':0.55,
                    'avail_bop_loss':0,
                    'avail_grid_loss':0,
                    'avail_turb_loss':0,
                    'elec_eff_loss':0,
                    'elec_parasitic_loss':0,
                    'env_degrad_loss':0,
                    'env_env_loss':0,
                    'env_icing_loss':0,
                    'ops_env_loss':0,
                    'ops_grid_loss':0,
                    'ops_load_loss':0,
                    'turb_generic_loss':0,
                    'turb_hysteresis_loss':0,
                    'turb_perf_loss':0,
                    'turb_specific_loss':0,
                    'wake_ext_loss':0}

# Set location info for Excel workbook with ATB values
filename = "2022 v3 Annual Technology Baseline Workbook Corrected 1-24-2023.xlsx"
sheets = {'NGCC':'Natural Gas_FE',
          'CCS': 'Natural Gas_FE',
          'PV':  'Solar - Utility PV',
          'LBW': 'Land-Based Wind',
          'OSW': 'Offshore Wind'}
base_cols =    {'NGCC':'M',
                'CCS': 'M',
                'PV':  'M',
                'LBW': 'M',
                'OSW': 'M'}
year_rows =    {'NGCC':68,
                'CCS': 68,
                'PV':  74,
                'LBW': 75,
                'OSW': 76}
# Rows with engineering/financial inputs (MAY NEED TO CHANGE IF HOPP/SAM CHANGES)
engin ={'NGCC':{'heat_rate_BTU_Wh':{'Advanced':72,
                                    'Moderate':73,
                                    'Conservative':74}},
        'CCS': {'heat_rate_BTU_Wh':{'Advanced':75,
                                    'Moderate':76,
                                    'Conservative':77}},
        'PV':  {'bifaciality':     {'Advanced':75,
                                    'Moderate':76,
                                    'Conservative':77},
                'albedo':          {'Advanced':78,
                                    'Moderate':79,
                                    'Conservative':80},
                'losses':          {'Advanced':81,
                                    'Moderate':82,
                                    'Conservative':83},
                'dc_degradation':  {'Advanced':84,
                                    'Moderate':85,
                                    'Conservative':86}},
        'LBW': {'hub_height':      {'Advanced':76,
                                    'Moderate':77,
                                    'Conservative':78},
                'rotor_diameter':  {'Advanced':79,
                                    'Moderate':80,
                                    'Conservative':81},
                'turbine_rating_kw':{'Advanced':82,
                                    'Moderate':83,
                                    'Conservative':84}},
        'OSW': {'hub_height':      {'Advanced':76,
                                    'Moderate':77,
                                    'Conservative':78},
                'rotor_diameter':  {'Advanced':79,
                                    'Moderate':80,
                                    'Conservative':81},
                'turbine_rating_kw':{'Advanced':82,
                                    'Moderate':83,
                                    'Conservative':84}}}
# OCC: Overnight capital cost ($/kW), FOM: Fixed O&M ($/kW-yr), VOM: Variable O&M ($/MWh)
finance =  {'NGCC':{'OCC_$_kW':    {'Advanced':105,
                                    'Moderate':106,
                                    'Conservative':107}},
            'CCS': {'OCC_$_kW':    {'Advanced':108,
                                    'Moderate':109,
                                    'Conservative':110}},
            'PV':  {'OCC_$_kW':    {'Advanced':203,
                                    'Moderate':204,
                                    'Conservative':205},
                    'FOM_$_kWyr':  {'Advanced':235,
                                    'Moderate':236,
                                    'Conservative':237},
                    'VOM_$_MWh':   {'Advanced':267,
                                    'Moderate':268,
                                    'Conservative':269}},
            'LBW': {'OCC_$_kW':    {'Advanced':204,
                                    'Moderate':205,
                                    'Conservative':206},
                    'FOM_$_kWyr':  {'Advanced':236,
                                    'Moderate':237,
                                    'Conservative':238},
                    'VOM_$_MWh':   {'Advanced':268,
                                    'Moderate':269,
                                    'Conservative':270}},
            'OSW': {'OCC_$_kW':    {'Advanced':253,
                                    'Moderate':254,
                                    'Conservative':255},
                    'FOM_$_kWyr':  {'Advanced':297,
                                    'Moderate':298,
                                    'Conservative':299},
                    'VOM_$_MWh':   {'Advanced':341,
                                    'Moderate':342,
                                    'Conservative':343}}}

# Import from spreadsheet
workbook = load_workbook(resource_dir/filename, read_only=True, data_only=True)
atb_tech = ['NGCC','CCS','PV','LBW','OSW']
for tech in atb_tech:
    sheet_name = sheets[tech]
    worksheet = workbook[sheet_name]
    # Figure out column range for years
    base_col = base_cols[tech]
    base_col_idx = column_index_from_string(base_col)
    year_row = str(year_rows[tech])
    base_year = worksheet[base_col+year_row].value
    start_col_idx = base_col_idx+sim_start_year-base_year
    end_col_idx = base_col_idx+sim_end_year-base_year
    start_col = get_column_letter(start_col_idx)
    end_col = get_column_letter(end_col_idx)
    # Replace row numbers in nested dicts with engin/finance params
    for param_dict in [engin, finance]:
        for param in param_dict[tech].values():
            for scenario, row in param.items():
                cells = worksheet[start_col+str(row)+':'+end_col+str(row)][0]
                values = []
                for i, cell in enumerate(cells):
                    value = cell.value
                    if param_dict is finance:
                        value = inflate(value, basis_year, sim_basis_year)
                    values.append(value)
                param[scenario] = values

#endregion

## TODO: Import natural gas price scenario
NG_price_MMBTU = [NG_price_MMBTU]*len(sim_years)

## TODO: Import natural gas price scenario
H2O_price_Tgal = [H2O_price_Tgal]*len(sim_years)

## Import the standard NETL NGCC power plant <NETL-PUB-22638/B31A & B31B>
#region

# Values in NETL report
NETL_basis_year = 2018
fullcap_MW =           {'NGCC': 727, 
                        'CCS':  646}
fullcap_C_kg_hr =      {'NGCC': 67890, 
                        'CCS':  61090}
fullcap_H2O_Tgal_day = {'NGCC': 2090, 
                        'CCS':  3437}
FOM_kW_yr =            {'NGCC': inflate(26.792,NETL_basis_year,basis_year),
                        'CCS':  inflate(63.911,NETL_basis_year,basis_year)}
VOM_other_MWh =        {'NGCC': inflate(1.70517-0.22769,NETL_basis_year,basis_year),
                        'CCS':  inflate(5.63373-0.42133,NETL_basis_year,basis_year)}
CO2_kg_MWh = {}
H2O_Tgal_MWh = {}

# Scale NGCC plant with CCS
NGCC_cap = {'NGCC': NGCC_cap,
            'CCS': NGCC_cap*fullcap_MW['CCS']/fullcap_MW['NGCC']}

# Convert to values per MWh produced
for key, value in fullcap_C_kg_hr.items():
    CO2_kg_MWh[key] = value*(C_MW+O_MW*2)/C_MW/fullcap_MW[key]
for key, value in fullcap_H2O_Tgal_day.items():
    H2O_Tgal_MWh[key] = value/24/fullcap_MW[key]

# Add items to engin nested dict
for plant in ['NGCC','CCS']:
    engin[plant]['full_cap_elec_out_MWh_yr'] = fullcap_MW[plant]*8760
    engin[plant]['elec_out_MWh_yr'] = NGCC_cap[plant]*NGCC_Cf*8760
    engin[plant]['output_kW'] = NGCC_cap[plant]*NGCC_Cf*1000
    engin[plant]['CO2_out_kg_MWh'] = CO2_kg_MWh[plant]
    engin[plant]['H2O_in_Tgal_MWh'] = H2O_Tgal_MWh[plant]

# Add items to finance nested dict
VOM_comps =['VOM_H2O_$_MWh',
            'VOM_NG_$_MWh',
            'VOM_other_$_MWh']
for plant in ['NGCC','CCS']:
    MWh_yr = engin[plant]['elec_out_MWh_yr']
    Tgal_MWh = engin[plant]['H2O_in_Tgal_MWh']
    finance[plant]['FOM_$_kWyr'] = FOM_kW_yr[plant]
    finance[plant]['FOM_$_yr'] = FOM_kW_yr[plant] * MWh_yr / 8760 * 1000
    finance[plant]['VOM_other_$_MWh'] = VOM_other_MWh[plant]
    finance[plant]['OCC_$'] = dict([(i,[]) for i in atb_scenarios])
    finance[plant]['VOM_H2O_$_MWh'] = dict([(i,[]) for i in atb_scenarios])
    finance[plant]['VOM_H2O_$_yr'] = dict([(i,[]) for i in atb_scenarios])
    finance[plant]['VOM_NG_$_MWh'] = dict([(i,[]) for i in atb_scenarios])
    finance[plant]['VOM_NG_$_yr'] = dict([(i,[]) for i in atb_scenarios])
    finance[plant]['VOM_$_MWh'] = dict([(i,[]) for i in atb_scenarios])
    finance[plant]['VOM_$_yr'] = dict([(i,[]) for i in atb_scenarios])
    for scenario in atb_scenarios:
        for i in range(len(sim_years)):
            price_Tgal = H2O_price_Tgal[i]
            price_MMBTU = NG_price_MMBTU[i]
            BTU_Wh = engin[plant]['heat_rate_BTU_Wh'][scenario][i]
            OCC_kW = finance[plant]['OCC_$_kW'][scenario][i]
            finance[plant]['OCC_$'][scenario].append(OCC_kW * MWh_yr / 8760 * 1000)
            finance[plant]['VOM_H2O_$_MWh'][scenario].append(price_Tgal*Tgal_MWh)
            finance[plant]['VOM_H2O_$_yr'][scenario].append(price_Tgal*Tgal_MWh*MWh_yr)
            finance[plant]['VOM_NG_$_MWh'][scenario].append(price_MMBTU*BTU_Wh)
            finance[plant]['VOM_NG_$_yr'][scenario].append(price_MMBTU*BTU_Wh*MWh_yr)
            VOM_MWh = 0
            for VOM_comp in VOM_comps:
                if type(finance[plant][VOM_comp]) is dict:
                    if type(finance[plant][VOM_comp][scenario]) is list:
                        VOM_MWh += finance[plant][VOM_comp][scenario][i]
                    else:
                        VOM_MWh += finance[plant][VOM_comp][scenario]
                else:
                    VOM_MWh += finance[plant][VOM_comp]
            finance[plant]['VOM_$_MWh'][scenario].append(VOM_MWh)
            finance[plant]['VOM_$_yr'][scenario].append(VOM_MWh*MWh_yr)

#endregion

# ## Import the NGCC plant locations and create list of survey locations
# #region

# earth_rad = 6371 # Earth's radium in km, needed for lat/long calcs

# # Loop through locations (begin as list of 1 NGCC plant location)
# for id, loc in locations.items():
#     on_land = loc['on_land'][0]
#     survey_rad = land_rad if on_land else osw_rad
#     lat = loc['lat'][0]
#     lon = loc['lon'][0]
    
#     # Add inner circle of 6 surrounding locations @ half of survey radius
#     in_circle_lat_delta = [3**.5/4, 0, -(3**.5/4), -(3**.5/4), 0,  3**.5/4]
#     for i, lat_delta in enumerate(in_circle_lat_delta):
#         lat_delta = survey_rad/earth_rad*lat_delta
#         lon_delta = ((survey_rad**2/4/earth_rad**2-lat_delta**2)/\
#                         (np.cos(lat/180*np.pi)**2))**.5*180/np.pi
#         lat_delta *= 180/np.pi
#         if i<3: lon_delta = -lon_delta 
#         loc['lat'].append(lat+lat_delta)
#         loc['lon'].append(lon+lon_delta)
#         loc['on_land'].append(globe.is_land(lat+lat_delta,lon+lon_delta))

#     # Add outer circle of 12 surrounding location @ full survey radius
#     out_circle_lat_delta = [ 1,   3**.5/2,   .5, 0, -.5, -(3**.5/2),
#                             -1, -(3**.5/2), -.5, 0,  .5,   3**.5/2]
#     for i, lat_delta in enumerate(out_circle_lat_delta):
#         lat_delta = survey_rad/earth_rad*lat_delta
#         lon_delta = ((survey_rad**2/earth_rad**2-lat_delta**2)/\
#                         (np.cos(lat/180*np.pi)**2))**.5*180/np.pi
#         lat_delta *= 180/np.pi
#         if i<6: lon_delta = -lon_delta 
#         loc['lat'].append(lat+lat_delta)
#         loc['lon'].append(lon+lon_delta)
#         loc['on_land'].append(globe.is_land(lat+lat_delta,lon+lon_delta))

# #endregion

# # Plot survey locations to check
# #region

# # Stolen from gis.stackexchange.com/questions/156035
# def merc_x(lon):
#   r_major=6378137.000
#   return r_major*(lon*np.pi/180)
# def merc_y(lat):
#   if lat>89.5:lat=89.5
#   if lat<-89.5:lat=-89.5
#   r_major=6378137.000
#   r_minor=6356752.3142
#   temp=r_minor/r_major
#   eccent=(1-temp**2)**.5
#   phi=(lat*np.pi/180)
#   sinphi=np.sin(phi)
#   con=eccent*sinphi
#   com=eccent/2
#   con=((1.0-con)/(1.0+con))**com
#   ts=np.tan((np.pi/2-phi)/2)/con
#   y=0-r_major*np.log(ts)
#   return y

# # Set up background image
# plt.clf
# bg_img = 'RCC search area new.png'
# img = plt.imread(resource_dir/bg_img)
# ax = plt.gca()
# min_x = merc_x(-100)
# max_x = merc_x(-67)
# min_y = merc_y(25)
# max_y = merc_y(46)
# ax.imshow(img, extent=[min_x, max_x, min_y, max_y])
# # Plot survey locations
# for id, loc in locations.items():
#     for n in range(len(loc['lat'])):
#         lat = loc['lat'][n]
#         lon = loc['lon'][n]
#         color = [1*loc['on_land'][n],0,0]
#         x = merc_x(lon)
#         y = merc_y(lat)
#         ax.plot(x,y,'.',color=color)
# plt.xlim([min_x,max_x])
# plt.ylim([min_y,max_y])
# plt.show()

# #endregion

## TODO: Import NGCC plant stream table from NETL report

## TODO: Import Wind/Solar/NGCC plant LCI tables from OpenLCA

## Fit curve to MeOH plant CAPEX data from IRENA report
#region

# IRENA report data <ISBN 978-92-9260-320-5>
capacity_kt_y =  [7,    90,    30,   440,  16.3, 50,  1800,  100]
capex_mil_kt_y = [3.19, 2.325, 1.15, 1.26, 0.98, 1.9, 0.235, 0.62]
basis_year = 2020
capex_mil_ky_y = [inflate(i, basis_year, sim_basis_year) for i in capex_mil_kt_y]
sources = ['Hank 2018', 'Mignard 2003', 'Clausen 2010', 'Perez-Fortes 2016',
            'Rivera-Tonoco 2016',' Belloti 2019', 'Nyari 2020', 'Szima 2018']

# Fit curve
def exp_curve(x, a, b): return a*x**(-b)
coeffs, _ = curve_fit(exp_curve, capacity_kt_y, capex_mil_kt_y)
a_cap2capex_mil_kt_y = coeffs[0]
b_cap2capex_mil_kt_y = coeffs[1]
finance['MeOH'] = {}
finance['MeOH']['capex_mil_kt_y_A'] = a_cap2capex_mil_kt_y
finance['MeOH']['capex_mil_kt_y_B'] = b_cap2capex_mil_kt_y

# # Plot to check
# log_x = np.linspace(np.log(np.min(capacity_kt_y)),np.log(np.max(capacity_kt_y)),100)
# plt.clf()
# plt.plot(np.power(np.e,log_x),
#         a_cap2capex_mil_kt_y*np.power(np.power(np.e,log_x),-b_cap2capex_mil_kt_y))
# plt.plot(capacity_kt_y,capex_mil_kt_y,'.')
# plt.xscale('log')
# plt.show()

#endregion

## Import MeOH ASPEN process model results to determine H2 requirements
# TODO: Replace static Nyari placeholder (doi:10.1016/j.jcou.2020.101166)
#           with executable Tan model fed by NGCC plant stack stream
# TODO: Euros-dollars function
#region

# Nyari FOM model - FCI = "Fixed capital investment"
fci_capex_ratio = 1/1.47
direct_labor_euro_person_year = 60000
direct_indirect_ratio = 0.3
workers_kt_y = 56/1800
o_and_m_fci_ratio = 0.015
insurance_fci_ratio = 0.005
tax_fci_ratio = 0.005
overhead_capex_ratio = fci_capex_ratio  *  (o_and_m_fci_ratio + \
                                            insurance_fci_ratio + \
                                            tax_fci_ratio)

# Convert labor rate
Nyari_basis_year = 2018
euro_dollar_ratio = 0.848 #irs.gov/individuals/international-taxpayers/
                            #yearly-average-currency-exchange-rates
direct_labor_person_year = direct_labor_euro_person_year/euro_dollar_ratio
labor_person_year = direct_labor_person_year*(1+direct_indirect_ratio)
labor_person_year = inflate(labor_person_year, Nyari_basis_year, basis_year)

# Nyari reactor performance
mass_ratio_CO2_MeOH = 1.397 # kg CO2 in needed per kg of MeOH produced
mass_ratio_H2_MeOH = 0.199 # kg H2 in needed per kg of MeOH produced
conv_eff_CO2 = 0.9837 # % of CO2 in converted to MeOH (rest comes out stack)
conv_eff_H2 = 1 # % of H2 in converted to MeOH
MeOH_purity = 0.999 # purity of MeOH produced (rest is assumed H2O)
cool_use_kWh_kg_MeOH = 0.81 # Cooling utility usage, kWh/kg MeOH produced
elec_use_kWh_kg_MeOH = 0.175 # Electricity usage, kWh/kg MeOH produced
CO2_emission = 0.023 # kg CO2 emitted directly per kg MeOH

#endregion

## Scale MeOH plant to NGCC output 
#region

engin['MeOH'] = {}

# Define NG plant feed for MeOH scenarios
engin['MeOH']['CO2 source'] = {}
for scenario in MeOH_scenarios:
    source = 'CCS' if 'CO2' in scenario else 'NGCC'
    engin['MeOH']['CO2 source'][scenario] = source

# Scale MeOH plant
CO2_kt_yr_scenarios = {}
MeOH_kt_yr = {}
for scenario in MeOH_scenarios:
    source = engin['MeOH']['CO2 source'][scenario]
    CO2_kt_yr = CO2_kg_MWh[source]*NGCC_cap[source]*NGCC_Cf*8760/1e6
    CO2_kt_yr_scenarios[scenario] = CO2_kt_yr
    MeOH_kt_yr[scenario] = CO2_kt_yr*mass_ratio_CO2_MeOH

# Add items to engin nested dict
engin['MeOH']['CO2_kg_yr_in'] = {}
engin['MeOH']['MeOH_kg_yr'] = {}
engin['MeOH']['output_kW'] = {}
for scenario in MeOH_scenarios:
    CO2_kg_yr = CO2_kt_yr_scenarios[scenario]*1e6
    MeOH_kg_yr = MeOH_kt_yr[scenario]*1e6
    engin['MeOH']['CO2_kg_yr_in'][scenario] = CO2_kg_yr
    engin['MeOH']['MeOH_kg_yr'][scenario] = MeOH_kg_yr
    engin['MeOH']['output_kW'][scenario] = MeOH_kg_yr/8760/3600*MeOH_LHV_MJ_kg/1000
    
# Add items to finance nested dict
finance_params =   ['OCC_$_kW','OCC_$',
                    'FOM_$_kWyr','FOM_$_yr']
for i in finance_params: finance['MeOH'][i] = {}
for scenario in MeOH_scenarios:
    # Capex calculation - from IRENA curve section above
    kg_yr = engin['MeOH']['MeOH_kg_yr'][scenario]
    kW = kg_yr * MeOH_LHV_MJ_kg * 1000 / 8760 / 3600
    kt_yr = engin['MeOH']['MeOH_kg_yr'][scenario]/1e6
    a = finance['MeOH']['capex_mil_kt_y_A']
    b = finance['MeOH']['capex_mil_kt_y_B']
    capex = a*kt_yr**(-b) * 1e6
    capex_kW = capex/kW
    finance['MeOH']['OCC_$_kW'][scenario] = capex_kW
    finance['MeOH']['OCC_$'][scenario] = capex
    # Fixed opex calculation - from Nyari model section above (for now)
    labor_yr = labor_person_year*workers_kt_y*kt_yr
    overhead_yr = overhead_capex_ratio*capex
    fom_yr = labor_yr + overhead_yr
    fom_yr_kW = fom_yr/kW
    finance['MeOH']['FOM_$_kWyr'][scenario] = fom_yr_kW
    finance['MeOH']['FOM_$_yr'][scenario] = fom_yr

#endregion

# Scale H2 plant to H2 needs and calculate financials with H2A model
#region

engin['H2'] = {}
finance['H2'] = {}

# Scale H2 plant
engin['H2']['H2_kg_yr'] = {}
engin['H2']['H2_kg_day'] = {}
engin['H2']['output_kW'] = {}
MeOH_scenario = plant_scenarios['MeOH']
for scenario in H2A_scenarios:
    MeOH_kg_yr = engin['MeOH']['MeOH_kg_yr'][MeOH_scenario]
    H2_kg_yr = MeOH_kg_yr*mass_ratio_H2_MeOH
    H2_kg_day = H2_kg_yr/365
    engin['H2']['H2_kg_yr'][scenario] = H2_kg_yr
    engin['H2']['H2_kg_day'][scenario] = H2_kg_day
    engin['H2']['output_kW'][scenario] = H2_kg_yr/8760/3600*H2_LHV_MJ_kg/1000

# Use H2A model to get costs, engin params to size hybrid plant
engin_params = ['water_use_kg_kgH2','H2O_in_Tgal_MWh','elec_use_kWh_kgH2','elec_in_kW']
finance_params = ['OCC_$','OCC_$_kW','FOM_$_yr','FOM_$_kWyr','VOM_H2O_$_MWh','VOM_H2O_$_yr']
for engin_param in engin_params:
    engin['H2'][engin_param] = dict([(i,[]) for i in H2A_scenarios])
for finance_param in finance_params:
    finance['H2'][finance_param] = dict([(i,[]) for i in H2A_scenarios])
for scenario in H2A_scenarios:
    H2_kg_day = engin['H2']['H2_kg_day'][scenario]
    output_kW = engin['H2']['output_kW'][scenario]
    for i, year in enumerate(sim_years):
        H2A_out = H2AModel_costs(H2_Cf,H2_kg_day,year,scenario)
        H2_basis_year, H2_capex, H2_fixed_om, kgH2O_kgH2, kWh_kgH2 = H2A_out
        H2_capex = inflate(H2_capex, H2_basis_year, basis_year)
        H2_fixed_om = inflate(H2_fixed_om, H2_basis_year, basis_year)
        finance['H2']['OCC_$'][scenario].append(H2_capex)
        finance['H2']['OCC_$_kW'][scenario].append(H2_capex/output_kW)
        finance['H2']['FOM_$_yr'][scenario].append(H2_fixed_om)
        finance['H2']['FOM_$_kWyr'][scenario].append(H2_fixed_om/output_kW)
        engin['H2']['water_use_kg_kgH2'][scenario].append(kgH2O_kgH2)
        engin['H2']['elec_use_kWh_kgH2'][scenario].append(kWh_kgH2)
        elec_in_kW = kWh_kgH2 * H2_kg_day / 24
        engin['H2']['elec_in_kW'][scenario].append(elec_in_kW)
        H2O_Tgal_MWh = kgH2O_kgH2 / kWh_kgH2 / L_gal
        engin['H2']['H2O_in_Tgal_MWh'][scenario].append(H2O_Tgal_MWh)
        H2O_VOM_MWh = H2O_Tgal_MWh * H2O_price_Tgal[i]
        finance['H2']['VOM_H2O_$_MWh'][scenario].append(H2O_VOM_MWh)
        MWh_yr = output_kW/1000*8760
        H2O_VOM_yr = H2O_VOM_MWh*MWh_yr
        finance['H2']['VOM_H2O_$_yr'][scenario].append(H2O_VOM_yr)

#endregion

## Write imported dicts to json dumpfiles

with open(Path(resource_dir/'engin.json'),'w') as file:
    json.dump(engin, file)
with open(Path(resource_dir/'finance.json'),'w') as file:
    json.dump(finance, file)
with open(Path(resource_dir/'scenario.json'),'w') as file:
    json.dump(scenario_info, file)

# %% Check imports and conversions

# Load imported dicts from json dumpfiles
import json
from pathlib import Path
resource_dir = Path(__file__).parent.absolute()/'resource_files'/'methanol_RCC'
with open(Path(resource_dir/'engin.json'),'r') as file:
    engin = json.load(file)
with open(Path(resource_dir/'finance.json'),'r') as file:
    finance = json.load(file)
with open(Path(resource_dir/'scenario.json'),'r') as file:
    scenario_info = json.load(file)
    plant_scenarios = scenario_info['plant_scenarios']
    atb_scenarios = scenario_info['atb_scenarios']
    H2A_scenarios = scenario_info['H2A_scenarios']
    MeOH_scenarios = scenario_info['MeOH_scenarios']

# Print prices per unit for ALL scenarios
prices_to_check = ['OCC_$_kW','FOM_$_kWyr','VOM_$_MWh']
plants_to_check = plant_scenarios.keys()
for plant in plants_to_check:
    if plant == 'H2':
        scenario_list = H2A_scenarios
    elif plant == 'MeOH':
        scenario_list = MeOH_scenarios
    else:
        scenario_list = atb_scenarios
    for scenario in scenario_list:
        for price in prices_to_check:
            try:
                price_value = finance[plant][price]
                if type(price_value) is dict:
                    price_value = price_value[scenario]
                    if type(price_value) is list:
                        price_value = price_value[0]
                print('{} plant {} ({} scenario): ${}'.format(plant,
                                                                price,
                                                                scenario,
                                                                price_value))
            except:
                print('{} plant did not import {} for {} scenario'.format(plant,
                                                                        price,
                                                                        scenario))

# Print absolute prices and parameters for specific scenarios
prices_to_check = ['OCC_$','FOM_$_yr','VOM_$_yr']
params_to_check = ['output_kW'] 
for plant in plants_to_check:
    scenario = plant_scenarios[plant]
    for price in prices_to_check:
        try:
            price_value = finance[plant][price]
            if type(price_value) is dict:
                price_value = price_value[scenario]
                if type(price_value) is list:
                    price_value = price_value[0]
            print('{} plant {} ({} scenario): ${}'.format(plant,
                                                            price,
                                                            scenario,
                                                            price_value))
        except:
            print('{} plant did not import {} for {} scenario'.format(plant,
                                                                    price,
                                                                    scenario))
    for param in params_to_check:
        try:
            param_value = engin[plant][param]
            if type(param_value) is dict:
                param_value = param_value[scenario]
                if type(param_value) is list:
                    param_value = param_value[0]
            print('{} plant {} ({} scenario): {}'.format(plant,
                                                            param,
                                                            scenario,
                                                            param_value))
        except:
            print('{} plant did not import {} for {} scenario'.format(plant,
                                                                    param,
                                                                    scenario))
# %%
