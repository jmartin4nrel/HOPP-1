'''
methanol_RCC.py

Generates a TEA for flue gas CO2 capture from an NGCC power plant and conversion
to methanol using hydrogen generated from a hybrid wind/solar plant nearby.
Uses NREL Annual Technology Baseline (ATB) future tech development scenarios.
Tracks all of the material streams and exports in a convenient format for LCA.

Author: Jonathan Martin <jonathan.martin@nrel.gov>
Uses ASPEN process model developed by Eric Tan <eric.tan@nrel.gov>
'''

## Import libraries needed
#region

import json
import copy
import numpy as np
import pandas as pd
from pathlib import Path
from global_land_mask import globe
import matplotlib.pyplot as plt
from scipy.optimize import curve_fit

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

from hybrid.sites import SiteInfo
from hybrid.hybrid_simulation import HybridSimulation
from hybrid.log import hybrid_logger as logger
from hybrid.keys import set_nrel_key_dot_env

from examples.H2_Analysis.H2AModel_capex_opex_scenarios import H2AModel_costs
from examples.methanol_RCC.extract_cambium_data import extract_cambium_data

#endregion

## Convert dollars to correct year
#region

def inflate(dollars, original_year, new_year):

    CPI = [172.2, 177.1, 179.9, 184.0, 188.9, 195.3, 201.6, 207.3, 215.3, 214.5,
           218.1, 224.9, 229.6, 233.0, 236.7, 237.0, 240.0, 245.1, 251.1, 255.7,
           258.8, 271.0, 292.7] # US City Average, all items, https://data.bls.gov/pdq/SurveyOutputServlet
    orig_cpi = CPI[int(original_year)-2000]
    new_cpi = CPI[int(new_year)-2000]
    dollars *= new_cpi/orig_cpi

    # standard_inflation = 0.025
    # dollars *= (1+standard_inflation)**(new_year-original_year)

    return dollars
    
#endregion    

def try_H2_ratio(H2_ratio=0.44, CO2_feed_mt_yr=1596153, ASPEN_MeOH_cap_mt_yr=115104, ASPEN_capex=33339802, ASPEN_Fopex=14.62, ASPEN_Vopex_cat=410.94, ASPEN_Vopex_other=-90.4, ASPEN_elec_use=0.89):

    ## ALL CHANGES HERE - SELECT SCENARIO COMBOS, MANUALLY OVERRIDE PRICING ASSUMPTIONS
    #region

    # Simulation duration
    sim_start_year = 2020
    sim_end_year = 2050
    sim_increment = 5
    sim_years = np.arange(sim_start_year,sim_end_year+sim_increment,sim_increment)

    # Financial constants
    sim_basis_year = 2020
    plant_lifespan = 30 # years
    discount_rate = 0.0707 # (= fixed charge rate)
    TASC_multiplier = 1.093 # total overnight cost * TASC multiplier = total as-spent cost

    # Commodity prices
    orig_NG_price_mmbtu = 4.56 # 2020 $/MMBTU TODO: make variable NG price scenarios
    H2O_price_tgal = 2.56 # 2020 $/Tgal
    CO2_TS_price_kg = 0.01092 # CO2 transmission and storage costs - DOE/NETL-2019/2044

    # Locations
    select_locations = False # TODO: Switch to True to only analyze locations listed below
    resource_dir = Path(__file__).parent.absolute()/'..'/'resource_files'/'methanol_RCC'
    location_file = 'ngcc_sites.csv'
    locations = {}
    states_covered = {}
    locations_df = pd.read_csv(resource_dir/location_file,index_col='PlantID')
    for location_num in range(locations_df.shape[0]):
        location = locations_df.iloc[location_num]
        state = location['PlantState']
        if state not in states_covered:
            states_covered[state] = 0
        states_covered[state] += 1
        locations[state+'{:02}'.format(states_covered[state])] = {'on_land':[True,],'lat':[location['Latitude'],],'lon':[location['Longitude'],]}
    site_choice = 'TX09' # Choose NGCC site to analyze all years/scenarios with HOPP
    site_num_choice = 11 # Choose site number (of 19 surrounding survey sites) to analyze
    min_plant_dist = 120 # km, minimum distance between NGCC plants in survey
    land_rad = 50 # km radius of survey area around NGCC plant on land
    osw_rad = 20 # km radium of survey area around offshore wind location

    # TODO: Override HOPP-calculated electricity and hydrogen prices
    manual_LCOE = False # Switch to True to override HOPP-calculated LCOE
    LCOE_kwh = 0.04 # $/kWh
    manual_LCOH = False # Switch to True to override HOPP-calculated LCOE
    LCOH_kg = 0.02 # $/kg

    # List possible dict definitions to use certain scenarios for different technologies
    atb_scenarios = ['Advanced','Moderate','Conservative']
    H2A_scenarios = ['Current','Future']
    MeOH_scenarios = ['Great','Good','OK']
    cambium_scenarios = ['MidCase','HighNGPrice','LowNGPrice']
    # Set specific scenarios to size with
    cambium_scenario = 'MidCase'
    plant_scenarios =  {'NGCC':'Advanced',
                        'CCS': 'Advanced',
                        'PV':  'Advanced',
                        'LBW': 'Advanced',
                        'OSW': 'Advanced',
                        'HCO2':'Future',
                        'HPSR':'Future',
                        'MSMR':'Great',
                        'MSMC':'Great',
                        'MCO2':'Great',
                        'MPSR':'Great'}
    scenario_info = {'plant_scenarios': plant_scenarios,
                    'cambium_scenario': cambium_scenario,
                    'atb_scenarios':    atb_scenarios,
                    'H2A_scenarios':    H2A_scenarios,
                    'MeOH_scenarios':   MeOH_scenarios,
                    'cambium_scenarios':cambium_scenarios,
                    'sim_years':        [int(i) for i in sim_years],
                    'site_selection':  {'site_name':site_choice,
                                        'site_num': site_num_choice}} 

    H2_Cf = 0.97 # capacity factor of H2 plants to scale results to
    Orig_NGCC_cap = 100 # MW, output of NGCC plant to scale results to
    NGCC_Cf = 0.85 # capacity factor of NGCC plant to scale results to
    ''' NGCC_cap*NGCC_Cf must be <170 MW for H2A model scaling to stay valid!
        (H2 output needs to stay below 200,000 kg H2/day'''
    MeOH_cap_mt_yr = ASPEN_MeOH_cap_mt_yr # Methanol capacity to scale results to, metric tons / yr

    cambium_dir = Path(__file__).parent.absolute()/'..'/'..'/'..'/'..'/'..'/'Projects'/'22 CO2 to Methanol'/'Cambium Data'
    
    #endregion

    ## Physical constants
    #region

    C_MW = 12.011 # g/mol C
    H_MW = 1.008 # g/mol H
    O_MW = 15.999 # g/mol O
    L_gal = 3.78541 # L/gal
    kg_lb = 0.453592 # kg/lb
    kJ_btu = 1.05506 # kJ/BTU

    MeOH_LHV_MJ_kg = 20.1 # Methanol net calorific value, MJ/kg
    NG_LHV_MJ_kg = 47.1 # Natural gas net calorific value, MJ/kg
    H2_LHV_MJ_kg = 120 # Hygrogen net calorific value, MJ/kg

    #endregion

    ## Import the NREL ATB scenario pricing info
    # TODO: Break out BOS costs (not in spreadsheets...)
    #region

    # Set location info for Excel workbook with ATB values
    filename = "2022 v3 Annual Technology Baseline Workbook Corrected 1-24-2023.xlsx"
    sheets = {'NGCC':'Natural Gas_FE',
            'CCS': 'Natural Gas_FE',
            'PV':  'Solar - Utility PV',
            'LBW': 'Land-Based Wind',
            'OSW': 'Offshore Wind'}
    base_cols =    {'NGCC':'M',
                    'CCS': 'M',
                    'PV':  'M',
                    'LBW': 'M',
                    'OSW': 'M'}
    year_rows =    {'NGCC':68,
                    'CCS': 68,
                    'PV':  74,
                    'LBW': 75,
                    'OSW': 76}
    # Rows with engineering/financial inputs (MAY NEED TO CHANGE IF HOPP/SAM CHANGES)
    engin ={'NGCC':{'heat_rate_btu_wh':{'Advanced':72,
                                        'Moderate':73,
                                        'Conservative':74}},
            'CCS': {'heat_rate_btu_wh':{'Advanced':75,
                                        'Moderate':76,
                                        'Conservative':77}},
            'PV':  {'bifaciality':     {'Advanced':75,
                                        'Moderate':76,
                                        'Conservative':77},
                    'albedo':          {'Advanced':78,
                                        'Moderate':79,
                                        'Conservative':80},
                    'losses':          {'Advanced':81,
                                        'Moderate':82,
                                        'Conservative':83},
                    'dc_degradation':  {'Advanced':84,
                                        'Moderate':85,
                                        'Conservative':86},
                    'inv_eff':         {'Advanced':87,
                                        'Moderate':88,
                                        'Conservative':89},
                    'capacity_factor': {'Advanced':90,
                                        'Moderate':91,
                                        'Conservative':92}},
            'LBW': {'hub_height':      {'Advanced':76,
                                        'Moderate':77,
                                        'Conservative':78},
                    'rotor_diameter':  {'Advanced':79,
                                        'Moderate':80,
                                        'Conservative':81},
                    'turbine_rating_kw':{'Advanced':82,
                                        'Moderate':83,
                                        'Conservative':84},
                    'capacity_factor': {'Advanced':85,
                                        'Moderate':86,
                                        'Conservative':87}},
            'OSW': {'hub_height':      {'Advanced':77,
                                        'Moderate':78,
                                        'Conservative':79},
                    'rotor_diameter':  {'Advanced':80,
                                        'Moderate':81,
                                        'Conservative':82},
                    'turbine_rating_kw':{'Advanced':83,
                                        'Moderate':84,
                                        'Conservative':85},
                    'capacity_factor': {'Advanced':86,
                                        'Moderate':87,
                                        'Conservative':88}}}
    # OCC: Overnight capital cost ($/kW), FOM: Fixed O&M ($/kW-yr), VOM: Variable O&M ($/MWh)
    finance =  {'NGCC':{'OCC_$_kw':    {'Advanced':105,
                                        'Moderate':106,
                                        'Conservative':107}},
                'CCS': {'OCC_$_kw':    {'Advanced':108,
                                        'Moderate':109,
                                        'Conservative':110}},
                'PV':  {'OCC_$_kw':    {'Advanced':203,
                                        'Moderate':204,
                                        'Conservative':205},
                        'FOM_$_kwyr':  {'Advanced':235,
                                        'Moderate':236,
                                        'Conservative':237},
                        'VOM_$_mwh':   {'Advanced':267,
                                        'Moderate':268,
                                        'Conservative':269}},
                'LBW': {'OCC_$_kw':    {'Advanced':204,
                                        'Moderate':205,
                                        'Conservative':206},
                        'FOM_$_kwyr':  {'Advanced':236,
                                        'Moderate':237,
                                        'Conservative':238},
                        'VOM_$_mwh':   {'Advanced':268,
                                        'Moderate':269,
                                        'Conservative':270}},
                'OSW': {'OCC_$_kw':    {'Advanced':253,
                                        'Moderate':254,
                                        'Conservative':255},
                        'FOM_$_kwyr':  {'Advanced':297,
                                        'Moderate':298,
                                        'Conservative':299},
                        'VOM_$_mwh':   {'Advanced':341,
                                        'Moderate':342,
                                        'Conservative':343}}}

    # Import from spreadsheet
    basis_year = 2020
    workbook = load_workbook(resource_dir/filename, read_only=True, data_only=True)
    atb_tech = ['NGCC','CCS','PV','LBW','OSW']
    for tech in atb_tech:
        sheet_name = sheets[tech]
        worksheet = workbook[sheet_name]
        # Figure out column range for years
        base_col = base_cols[tech]
        base_col_idx = column_index_from_string(base_col)
        year_row = str(year_rows[tech])
        base_year = worksheet[base_col+year_row].value
        start_col_idx = base_col_idx+sim_start_year-base_year
        end_col_idx = base_col_idx+sim_end_year-base_year
        start_col = get_column_letter(start_col_idx)
        end_col = get_column_letter(end_col_idx)
        # Replace row numbers in nested dicts with engin/finance params
        for param_dict in [engin, finance]:
            for param in param_dict[tech].values():
                for scenario, row in param.items():
                    cells = worksheet[start_col+str(row)+':'+end_col+str(row)][0]
                    values = []
                    for i, cell in enumerate(cells):
                        if i+base_year in sim_years:
                            value = cell.value
                            if param_dict is finance:
                                value = inflate(value, basis_year, sim_basis_year)
                            values.append(value)
                    param[scenario] = values

    # Set constants not given in spreadsheet

    bos_ratio = {'LBW': 322/1462, 
                'PV': 20/89}
    for tech in bos_ratio.keys():
        finance[tech]['OCC_bos_ratio'] = {}
        for scenario in atb_scenarios:
            finance[tech]['OCC_bos_ratio'][scenario] = bos_ratio[tech]

    pv_constants = {'array_type':2,
                    'azimuth':180,
                    'dc_ac_ratio':1.28} # ILR NOT OPTIMIZED - can HOPP do this?
    for constant in pv_constants:
        engin['PV'][constant] = {}
        for scenario in atb_scenarios:
            engin['PV'][constant][scenario] = pv_constants[constant]

    wind_constants =   {'wind_turbine_max_cp':0.45,
                        'avail_bop_loss':0,
                        'avail_grid_loss':0,
                        'avail_turb_loss':0,
                        'elec_eff_loss':0,
                        'elec_parasitic_loss':0,
                        'env_degrad_loss':0,
                        'env_env_loss':0,
                        'env_icing_loss':0,
                        'ops_env_loss':0,
                        'ops_grid_loss':0,
                        'ops_load_loss':0,
                        'turb_generic_loss':0,
                        'turb_hysteresis_loss':0,
                        'turb_perf_loss':0,
                        'turb_specific_loss':0,
                        'wake_ext_loss':0}
    for constant in wind_constants:
        engin['LBW'][constant] = {}
        engin['OSW'][constant] = {}
        for scenario in atb_scenarios:
            engin['LBW'][constant][scenario] = wind_constants[constant]
            engin['OSW'][constant][scenario] = wind_constants[constant]

    #endregion

    ## TODO: Import natural gas price scenario
    NG_price_mmbtu = {}
    extract_cambium_data(cambium_dir, basis_year)
    with open(Path(resource_dir/('cambium_ng_'+cambium_scenario+'.json')),'r') as file:
        cambium_ng = pd.read_json(file)
    for scenario in cambium_scenarios:
        NG_price_mmbtu[scenario] = [cambium_ng.loc[year].values[0] for year in sim_years]

    ## TODO: Import water price scenario
    H2O_price_tgal = [H2O_price_tgal]*len(sim_years)

    ## Fit curve to MeOH plant CAPEX data from IRENA report
    #region

    # IRENA report data <ISBN 978-92-9260-320-5>
    capacity_kt_y =  [7,    90,    30,   440,  16.3, 50,  1800,  100]
    capex_mil_kt_y = [3.19, 2.325, 1.15, 1.26, 0.98, 1.9, 0.235, 0.62]
    basis_year = 2020
    capex_mil_kt_y = [inflate(i, basis_year, sim_basis_year) for i in capex_mil_kt_y]
    sources = ['Hank 2018', 'Mignard 2003', 'Clausen 2010', 'Perez-Fortes 2016',
                'Rivera-Tonoco 2016',' Belloti 2019', 'Nyari 2020', 'Szima 2018']

    # Fit curve
    def exp_curve(x, a, b): return a*x**(-b)
    coeffs, _ = curve_fit(exp_curve, capacity_kt_y, capex_mil_kt_y)
    a_cap2capex_mil_kt_y = coeffs[0]
    b_cap2capex_mil_kt_y = coeffs[1]
    finance['MCO2'] = {}
    finance['MCO2']['capex_mil_kt_y_A'] = a_cap2capex_mil_kt_y
    finance['MCO2']['capex_mil_kt_y_B'] = b_cap2capex_mil_kt_y
    finance['MPSR'] = {}
    finance['MPSR']['capex_mil_kt_y_A'] = a_cap2capex_mil_kt_y
    finance['MPSR']['capex_mil_kt_y_B'] = b_cap2capex_mil_kt_y
    # Add other methanol plants to finance matrix
    finance['MSMR'] = {}
    finance['MSMC'] = {}

    # # Plot to check
    # log_x = np.linspace(np.log(np.min(capacity_kt_y)),np.log(np.max(capacity_kt_y)),100)
    # plt.clf()
    # plt.plot(np.power(np.e,log_x),
    #         a_cap2capex_mil_kt_y*np.power(np.power(np.e,log_x),-b_cap2capex_mil_kt_y))
    # plt.plot(capacity_kt_y,capex_mil_kt_y,'.')
    # plt.xscale('log')
    # plt.show()

    #endregion

    ## Import MeOH ASPEN process model results to determine H2 requirements
    # TODO: Replace static Nyari placeholder (doi:10.1016/j.jcou.2020.101166)
    #           with executable Tan model fed by NGCC plant stack stream
    # TODO: Euros-dollars function
    #region

    # Nyari FOM model - FCI = "Fixed capital investment"
    fci_capex_ratio = 1/1.47
    direct_labor_euro_person_year = 60000
    direct_indirect_ratio = 0.3
    workers_kt_y = 56/1800
    o_and_m_fci_ratio = 0.015
    insurance_fci_ratio = 0.005
    tax_fci_ratio = 0.005
    overhead_capex_ratio = fci_capex_ratio  *  (o_and_m_fci_ratio + \
                                                insurance_fci_ratio + \
                                                tax_fci_ratio)

    # Convert labor rate
    Nyari_basis_year = 2018
    euro_dollar_ratio = 0.848 #irs.gov/individuals/international-taxpayers/
                                #yearly-average-currency-exchange-rates
    direct_labor_person_year = direct_labor_euro_person_year/euro_dollar_ratio
    labor_person_year = direct_labor_person_year*(1+direct_indirect_ratio)
    labor_person_year = inflate(labor_person_year, Nyari_basis_year, basis_year)

    # Nyari reactor performance
    nyari_mass_ratio_CO2_MeOH = 1.423#1.397 # kg CO2 in needed per kg of MeOH produced
    nyari_mass_ratio_H2_MeOH = 0.195#0.192 # kg H2 in needed per kg of MeOH produced
    nyari_CO2_conv_pct = 96.49#98.37 # pct CO2 converted
    nyari_elec_usage = 0.172 # kWh/kg MeOH

    # Nyari H2O/catalyst usage - ESTIMATED (reported in Euros/year, euros/kg)
    nyari_H2O_gal_kg_MeOH = 0.3
    nyari_cat_mg_kg_MeOH = 25
    nyari_cat_price_kg = 10

    # Ruddy reactor performance
    ruddy_mass_ratio_CO2_MeOH = {'Great':   CO2_feed_mt_yr/MeOH_cap_mt_yr,
                                'Good':    1.6,
                                'OK':      1.8} # kg CO2 in needed per kg of MeOH produced
    ruddy_mass_ratio_H2_MeOH =  {'Great':   H2_ratio,
                                'Good':    0.25,
                                'OK':      0.3} # kg H2 in needed per kg of MeOH produced
    ruddy_CO2_conv_pct =   {'Great':    MeOH_cap_mt_yr/CO2_feed_mt_yr*(C_MW+O_MW*2)/(C_MW+O_MW+H_MW*4)*100,
                            'Good':     95,
                            'OK':       90} # % CO2 converted to MeOH
    ruddy_elec_usage =     {'Great':    ASPEN_elec_use,
                            'Good':     0.15,
                            'OK':       0.2} # kWh/kg MeOH
    ruddy_cat_mg_kg_MeOH = {'Great':    25,
                            'Good':     30,
                            'OK':       35} # mg catalyst / kg MeOH 
    ruddy_cat_price_kg =   {'Great':    10,
                            'Good':     10,
                            'OK':       10} # $/kg catalyst 

    #endregion

    ## Scale MeOH plants
    #region

    for plant in ['MSMR','MSMC','MCO2','MPSR']:
        engin[plant] = {}
        engin[plant]['MeOH_LHV_MJ_kg'] = MeOH_LHV_MJ_kg

        # Define NG plant feed for MeOH scenarios
        engin[plant]['CO2 source'] = {}
        source = 'CCS' if 'CO2' in plant else 'NGCC'
        engin[plant]['CO2 source'] = source

        # Scale MeOH plant
        MeOH_kt_yr = {}
        CO2_kt_yr = {}
        H2_kt_yr = {}
        Stack_CO2_kt_yr = {}
        H2O_gal_day = {}
        elec_in_kw = {}
        cat_mg_kg_MeOH = {}
        if 'MS' not in plant:
            source = engin[plant]['CO2 source']
            for scenario in MeOH_scenarios:
                if 'CO2' in plant:
                    MeOH_kt_yr[scenario] = MeOH_cap_mt_yr/1e3
                    CO2_kt_yr[scenario] = nyari_mass_ratio_CO2_MeOH*MeOH_kt_yr[scenario]
                    H2_kt_yr[scenario] = MeOH_kt_yr[scenario]*nyari_mass_ratio_H2_MeOH
                    Stack_CO2_kt_yr[scenario] = MeOH_kt_yr[scenario]*(100-nyari_CO2_conv_pct)/100
                    elec_in_kw[scenario] = MeOH_kt_yr[scenario]*1e6/8760 * nyari_elec_usage
                    cat_mg_kg_MeOH[scenario] = nyari_cat_mg_kg_MeOH
                else:
                    MeOH_kt_yr[scenario] = MeOH_cap_mt_yr/1e3
                    CO2_kt_yr[scenario] = ruddy_mass_ratio_CO2_MeOH[scenario]*MeOH_kt_yr[scenario]
                    H2_kt_yr[scenario] = MeOH_kt_yr[scenario]*ruddy_mass_ratio_H2_MeOH[scenario]
                    Stack_CO2_kt_yr[scenario] = MeOH_kt_yr[scenario]*(100-ruddy_CO2_conv_pct[scenario])/100
                    elec_in_kw[scenario] = MeOH_kt_yr[scenario]*1e6/8760 * ruddy_elec_usage[scenario]
                    cat_mg_kg_MeOH[scenario] = ruddy_cat_mg_kg_MeOH[scenario]
                H2O_gal_day[scenario] = MeOH_kt_yr[scenario] * 1e6 / 365 * nyari_H2O_gal_kg_MeOH
            NG_MMBtu_day = 0
            TS_CO2_kt_yr = 0
            elec_out_kw = 0
        else:
            MeOH_lb_hr = 940989 # Source number from DOE/NETL-341/101514
            CO2_lb_hr_out = 235808 # Source number from DOE/NETL-341/101514
            SMR_Cf = 0.9 # Capacity factor, Source number from DOE/NETL-341/101514
            NG_MMBtu_day = 315872
            elec_out_kw = 2758/24*1e3
            cat_density_kg_m3 = 1500 # Taken from Nyari et al, just needed to get catalyst in kg
            m3_ft3 = (25.4*12/1000)**3
            cat_ft3_day = 3.72
            for scenario in MeOH_scenarios:
                H2_kt_yr[scenario] = 0
                CO2_kt_yr[scenario] = 0
                H2O_gal_day[scenario] = 2454
                CO2_kt_yr_out = CO2_lb_hr_out*8760*SMR_Cf*kg_lb/1e6
                MeOH_kt_yr[scenario] = MeOH_lb_hr*8760*SMR_Cf*kg_lb/1e6
                cat_mg_kg_MeOH[scenario] = cat_ft3_day*m3_ft3*cat_density_kg_m3*365/MeOH_kt_yr[scenario]
                elec_in_kw[scenario] = 0
                if 'SMC' in plant:
                    TS_CO2_kt_yr = CO2_kt_yr_out
                    Stack_CO2_kt_yr[scenario] = 0
                else:
                    TS_CO2_kt_yr = 0
                    Stack_CO2_kt_yr[scenario] = CO2_kt_yr_out

        # Add items to engin nested dict
        engin[plant]['CO2_kg_yr_in'] = {}
        engin[plant]['H2_kg_yr_in'] = {}
        engin[plant]['H2O_kg_yr_in'] = {}
        engin[plant]['NG_kg_yr_in'] = {}
        engin[plant]['Stack_CO2_kg_yr'] = {}
        engin[plant]['TS_CO2_kg_yr'] = {}
        engin[plant]['cat_kg_yr'] = {}
        engin[plant]['MeOH_kg_yr'] = {}
        engin[plant]['elec_in_kw'] = {}
        engin[plant]['elec_out_kw'] = {}
        engin[plant]['elec_in_mwh_yr'] = {}
        engin[plant]['elec_out_mwh_yr'] = {}
        engin[plant]['output_kw'] = {}
        for scenario in MeOH_scenarios:
            CO2_kg_yr = CO2_kt_yr[scenario]*1e6
            H2_kg_yr = H2_kt_yr[scenario]*1e6
            MeOH_kg_yr = MeOH_kt_yr[scenario]*1e6
            H2O_in_kg_yr = H2O_gal_day[scenario]*L_gal*365
            NG_in_kg_yr = NG_MMBtu_day*kJ_btu*365*1000/NG_LHV_MJ_kg
            Stack_CO2_kg_yr = Stack_CO2_kt_yr[scenario]*1e6
            TS_CO2_kg_yr = TS_CO2_kt_yr*1e6
            cat_kg_yr = cat_mg_kg_MeOH[scenario]*MeOH_kg_yr/1e6
            engin[plant]['CO2_kg_yr_in'][scenario] = CO2_kg_yr
            engin[plant]['H2_kg_yr_in'][scenario] = H2_kg_yr
            engin[plant]['MeOH_kg_yr'][scenario] = MeOH_kg_yr
            engin[plant]['H2O_kg_yr_in'][scenario] = H2O_in_kg_yr
            engin[plant]['NG_kg_yr_in'][scenario] = NG_in_kg_yr
            engin[plant]['TS_CO2_kg_yr'][scenario] = TS_CO2_kg_yr
            engin[plant]['Stack_CO2_kg_yr'][scenario] = Stack_CO2_kg_yr
            engin[plant]['cat_kg_yr'][scenario] = cat_kg_yr
            engin[plant]['elec_in_kw'][scenario] = elec_in_kw[scenario]
            engin[plant]['elec_in_mwh_yr'][scenario] = elec_in_kw[scenario]*8.76
            engin[plant]['elec_out_kw'][scenario] = elec_out_kw
            engin[plant]['elec_out_mwh_yr'][scenario] = elec_out_kw*8.76
            engin[plant]['output_kw'][scenario] = MeOH_kg_yr/8760/3600*MeOH_LHV_MJ_kg*1000
            
        # Add items to finance nested dict
        finance_params =   ['OCC_$_kw','OCC_$',
                            'FOM_$_kwyr','FOM_$_yr',
                            'VOM_other_$_mwh','VOM_other_$_yr',
                            'VOM_H2O_$_mwh','VOM_H2O_$_yr',
                            'VOM_NG_$_mwh','VOM_NG_$_yr',
                            'VOM_cat_$_mwh','VOM_cat_$_yr',
                            'VOM_TS_$_mwh','VOM_TS_$_yr']
        for i in finance_params: finance[plant][i] = {}
        for scenario in MeOH_scenarios:
            kg_yr = engin[plant]['MeOH_kg_yr'][scenario]
            kw = kg_yr * MeOH_LHV_MJ_kg * 1000 / 8760 / 3600
            mwh_yr = kw*8.76
            kt_yr = kg_yr/1e6
            # Capex calculation - from IRENA curve section above OR DOE/NETL-341-101541
            if 'MS' in plant:
                # Direct numbers from DOE/NETL-341-101541, page 57/60
                toc_tc = 2644295/2171740 # Total overnight cost to total cost ratio
                tc_base = 2007
                foc_voc_base = 2011
                if 'SMC' in plant:
                    tc = 2171740*1000
                    vom_other_yr = inflate(31389888+5371231-1349340,foc_voc_base,sim_basis_year)
                else:
                    tc = (2171740-299998)*1000
                    vom_other_yr = inflate(31389888+5371231-1349340-378799-192014,foc_voc_base,sim_basis_year)
                tc = inflate(tc,tc_base,sim_basis_year)
                capex = tc*toc_tc
                fom_yr = inflate(75244327,foc_voc_base,sim_basis_year)
                vom_other_mwh = vom_other_yr/mwh_yr
            elif 'MC' in plant:
                # From Nyari et al model
                a = finance[plant]['capex_mil_kt_y_A']
                b = finance[plant]['capex_mil_kt_y_B']
                capex = 1.5e6*kt_yr#a*kt_yr**(-b) * kt_yr * 1e6
                labor_yr = labor_person_year*workers_kt_y*kt_yr
                overhead_yr = overhead_capex_ratio*capex
                fom_yr = labor_yr + overhead_yr
                vom_other_yr = 0
                vom_other_mwh = 0
            else:
                # From ASPEN model
                capex = ASPEN_capex
                vom_other_mt = ASPEN_Vopex_other
                fom_yr = ASPEN_Fopex*kt_yr*1000
                vom_other_yr = vom_other_mt*kt_yr*1000
                vom_other_mwh = vom_other_yr/mwh_yr
            capex_kw = capex/kw
            fom_yr_kw = fom_yr/kw
            vom_h2o_yr = [engin[plant]['H2O_kg_yr_in'][scenario]*i/L_gal/1000 for i in H2O_price_tgal]
            vom_h2o_mwh = [i/mwh_yr for i in vom_h2o_yr]
            vom_cat_yr = engin[plant]['cat_kg_yr'][scenario]
            vom_cat_mwh = vom_cat_yr/mwh_yr
            vom_ts_yr = engin[plant]['TS_CO2_kg_yr'][scenario]*CO2_TS_price_kg
            vom_ts_mwh = vom_ts_yr/mwh_yr
            ng_kg_yr = engin[plant]['NG_kg_yr_in'][scenario]
            vom_ng_yr = []
            offset = sim_years[0]-2020
            for i in range(len(sim_years)):
                lcon = 0
                for j in range(plant_lifespan):
                    if j < len(NG_price_mmbtu[cambium_scenario][i+offset:]):
                        lcon += NG_price_mmbtu[cambium_scenario][i+offset+j]
                    else:
                        lcon += NG_price_mmbtu[cambium_scenario][-1]
                lcon /= plant_lifespan
                vom_ng_yr.append(ng_kg_yr*lcon/kJ_btu/1000*NG_LHV_MJ_kg)
            vom_ng_mwh = [i/mwh_yr for i in vom_ng_yr]
            if 'MP' in plant:
                vom_ng_mwh = 0
                vom_ng_yr = 0
                vom_h2o_mwh = 0
                vom_h2o_yr = 0
                vom_cat_yr = ASPEN_Vopex_cat*kt_yr*1000
                vom_cat_mwh = vom_cat_yr/mwh_yr
                vom_ts_mwh = 0
                vom_ts_yr = 0
            finance[plant]['OCC_$_kw'][scenario] = capex_kw
            finance[plant]['OCC_$'][scenario] = capex
            finance[plant]['FOM_$_kwyr'][scenario] = fom_yr_kw
            finance[plant]['FOM_$_yr'][scenario] = fom_yr
            finance[plant]['VOM_other_$_mwh'][scenario] = vom_other_mwh
            finance[plant]['VOM_other_$_yr'][scenario] = vom_other_yr
            finance[plant]['VOM_NG_$_mwh'][scenario] = vom_ng_mwh
            finance[plant]['VOM_NG_$_yr'][scenario] = vom_ng_yr
            finance[plant]['VOM_H2O_$_mwh'][scenario] = vom_h2o_mwh
            finance[plant]['VOM_H2O_$_yr'][scenario] = vom_h2o_yr
            finance[plant]['VOM_cat_$_mwh'][scenario] = vom_cat_mwh
            finance[plant]['VOM_cat_$_yr'][scenario] = vom_cat_yr
            finance[plant]['VOM_TS_$_mwh'][scenario] = vom_ts_mwh
            finance[plant]['VOM_TS_$_yr'][scenario] = vom_ts_yr

    #endregion

    # Scale H2 plants to H2 needs and calculate financials with H2A model
    #region

    for plant in ['HCO2','HPSR']:
        engin[plant] = {}
        engin[plant]['H2_LHV_MJ_kg'] = H2_LHV_MJ_kg
        finance[plant] = {}

        # Scale H2 plant
        engin[plant]['H2_kg_yr'] = {}
        engin[plant]['H2_kg_day'] = {}
        engin[plant]['output_kw'] = {}
        MeOH_scenario = plant_scenarios['MPSR']
        for scenario in H2A_scenarios:
            Mplant = 'M'+plant[1:]
            MeOH_kg_yr = engin[Mplant]['MeOH_kg_yr'][MeOH_scenario]
            H2_kg_yr = engin[Mplant]['H2_kg_yr_in'][MeOH_scenario]
            H2_kg_day = H2_kg_yr/365
            engin[plant]['H2_kg_yr'][scenario] = H2_kg_yr
            engin[plant]['H2_kg_day'][scenario] = H2_kg_day
            engin[plant]['output_kw'][scenario] = H2_kg_yr/8760/3600*H2_LHV_MJ_kg*1000

        # Use H2A model to get costs, engin params to size hybrid plant
        engin_params = ['water_use_kg_kgH2','H2O_in_tgal_mwh','elec_use_kwh_kgH2','elec_in_kw','elec_in_mwh_yr']
        finance_params = ['OCC_$','OCC_$_kw','FOM_$_yr','FOM_$_kwyr','VOM_H2O_$_mwh','VOM_H2O_$_yr']
        for engin_param in engin_params:
            engin[plant][engin_param] = dict([(i,[]) for i in H2A_scenarios])
        for finance_param in finance_params:
            finance[plant][finance_param] = dict([(i,[]) for i in H2A_scenarios])
        for scenario in H2A_scenarios:
            H2_kg_day = engin[plant]['H2_kg_day'][scenario]
            output_kw = engin[plant]['output_kw'][scenario]
            for i, year in enumerate(sim_years):
                H2A_out = H2AModel_costs(H2_Cf,H2_kg_day,year,scenario)
                H2_basis_year, H2_capex, H2_fixed_om, kgH2O_kgH2, kwh_kgH2 = H2A_out
                H2_capex = inflate(H2_capex, H2_basis_year, basis_year)
                H2_fixed_om = inflate(H2_fixed_om, H2_basis_year, basis_year)
                finance[plant]['OCC_$'][scenario].append(H2_capex)
                finance[plant]['OCC_$_kw'][scenario].append(H2_capex/output_kw)
                finance[plant]['FOM_$_yr'][scenario].append(H2_fixed_om)
                finance[plant]['FOM_$_kwyr'][scenario].append(H2_fixed_om/output_kw)
                engin[plant]['water_use_kg_kgH2'][scenario].append(kgH2O_kgH2)
                engin[plant]['elec_use_kwh_kgH2'][scenario].append(kwh_kgH2)
                elec_in_kw = kwh_kgH2 * H2_kg_day / 24
                engin[plant]['elec_in_kw'][scenario].append(elec_in_kw)
                engin[plant]['elec_in_mwh_yr'][scenario].append(elec_in_kw*8.76)
                H2O_tgal_mwh = kgH2O_kgH2 / kwh_kgH2 / L_gal
                engin[plant]['H2O_in_tgal_mwh'][scenario].append(H2O_tgal_mwh)
                H2O_VOM_mwh = H2O_tgal_mwh * H2O_price_tgal[i]
                finance[plant]['VOM_H2O_$_mwh'][scenario].append(H2O_VOM_mwh)
                mwh_yr = output_kw/1000*8760
                H2O_VOM_yr = H2O_VOM_mwh*mwh_yr
                finance[plant]['VOM_H2O_$_yr'][scenario].append(H2O_VOM_yr)

    #endregion


    ## Import the standard NETL NGCC power plant <NETL-PUB-22638/B31A & B31B>
    #region

    # Values in NETL report
    NETL_basis_year = 2018
    fullcap_mw =           {'NGCC': 727, 
                            'CCS':  646}
    fullcap_C_kg_hr =      {'NGCC': 67890, 
                            'CCS':  61090}
    fullcap_H2O_tgal_day = {'NGCC': 2090, 
                            'CCS':  3437}
    FOM_kw_yr =            {'NGCC': inflate(26.792,NETL_basis_year,basis_year),
                            'CCS':  inflate(63.911,NETL_basis_year,basis_year)}
    VOM_other_mwh =        {'NGCC': inflate(1.70517-0.22769,NETL_basis_year,basis_year),
                            'CCS':  inflate(5.63373-0.42133,NETL_basis_year,basis_year)}
    CO2_kg_mwh = {}
    H2O_tgal_mwh = {}

    # Scale NGCC plant with CCS
    NGCC_cap = {'NGCC': Orig_NGCC_cap,
                'CCS': Orig_NGCC_cap*fullcap_mw['CCS']/fullcap_mw['NGCC']}

    # Convert to values per MWh produced
    for key, value in fullcap_C_kg_hr.items():
        CO2_kg_mwh[key] = value*(C_MW+O_MW*2)/C_MW/fullcap_mw[key]
    for key, value in fullcap_H2O_tgal_day.items():
        H2O_tgal_mwh[key] = value/24/fullcap_mw[key]

    # Add items to engin nested dict
    for plant in ['NGCC','CCS']:
        engin[plant]['full_cap_elec_out_mwh_yr'] = fullcap_mw[plant]*8760
        engin[plant]['elec_out_mwh_yr'] = NGCC_cap[plant]*NGCC_Cf*8760
        engin[plant]['output_kw'] = NGCC_cap[plant]*NGCC_Cf*1000
        engin[plant]['CO2_out_kg_mwh'] = CO2_kg_mwh[plant]
        engin[plant]['H2O_in_tgal_mwh'] = H2O_tgal_mwh[plant]
        engin[plant]['capacity_factor'] = NGCC_Cf

    # Add items to finance nested dict
    VOM_comps =['VOM_H2O_$_mwh',
                'VOM_NG_$_mwh',
                'VOM_other_$_mwh']
    for plant in ['NGCC','CCS']:
        mwh_yr = engin[plant]['elec_out_mwh_yr']
        tgal_mwh = engin[plant]['H2O_in_tgal_mwh']
        finance[plant]['FOM_$_kwyr'] = FOM_kw_yr[plant]
        finance[plant]['FOM_$_yr'] = FOM_kw_yr[plant] * mwh_yr / NGCC_Cf / 8760 * 1000
        finance[plant]['VOM_other_$_mwh'] = VOM_other_mwh[plant]
        finance[plant]['OCC_$'] = dict([(i,[]) for i in atb_scenarios])
        finance[plant]['VOM_H2O_$_mwh'] = dict([(i,[]) for i in atb_scenarios])
        finance[plant]['VOM_H2O_$_yr'] = dict([(i,[]) for i in atb_scenarios])
        finance[plant]['VOM_NG_$_mwh'] = dict([(i,[]) for i in atb_scenarios])
        finance[plant]['VOM_NG_$_yr'] = dict([(i,[]) for i in atb_scenarios])
        finance[plant]['VOM_$_mwh'] = dict([(i,[]) for i in atb_scenarios])
        finance[plant]['VOM_$_yr'] = dict([(i,[]) for i in atb_scenarios])
        engin[plant]['NG_kg_yr_in'] = dict([(i,[]) for i in atb_scenarios])
        for scenario in atb_scenarios:
            for i in range(len(sim_years)):
                price_tgal = H2O_price_tgal[i]
                offset = sim_years[0]-2020
                price_mmbtu = 0
                for j in range(plant_lifespan):
                    if j < len(NG_price_mmbtu[cambium_scenario][i+offset:]):
                        price_mmbtu += NG_price_mmbtu[cambium_scenario][i+offset+j]
                    else:
                        price_mmbtu += NG_price_mmbtu[cambium_scenario][-1]
                price_mmbtu /= plant_lifespan
                btu_wh = engin[plant]['heat_rate_btu_wh'][scenario][i]
                ng_kg_yr = btu_wh*mwh_yr*kJ_btu*1000/NG_LHV_MJ_kg
                OCC_kw = finance[plant]['OCC_$_kw'][scenario][i]
                finance[plant]['OCC_$'][scenario].append(OCC_kw * mwh_yr / NGCC_Cf / 8760 * 1000)
                finance[plant]['VOM_H2O_$_mwh'][scenario].append(price_tgal*tgal_mwh)
                finance[plant]['VOM_H2O_$_yr'][scenario].append(price_tgal*tgal_mwh*mwh_yr)
                finance[plant]['VOM_NG_$_mwh'][scenario].append(price_mmbtu*btu_wh)
                finance[plant]['VOM_NG_$_yr'][scenario].append(price_mmbtu*btu_wh*mwh_yr)
                engin[plant]['NG_kg_yr_in'][scenario].append(ng_kg_yr)
                VOM_mwh = 0
                for VOM_comp in VOM_comps:
                    if type(finance[plant][VOM_comp]) is dict:
                        if type(finance[plant][VOM_comp][scenario]) is list:
                            VOM_mwh += finance[plant][VOM_comp][scenario][i]
                        else:
                            VOM_mwh += finance[plant][VOM_comp][scenario]
                    else:
                        VOM_mwh += finance[plant][VOM_comp]
                finance[plant]['VOM_$_mwh'][scenario].append(VOM_mwh)
                finance[plant]['VOM_$_yr'][scenario].append(VOM_mwh*mwh_yr)

    #endregion

    ## Import the NGCC plant locations and create list of survey locations
    #TODO: Actually import plant locations from file
    #region

    earth_rad = 6371 # Earth's radium in km, needed for lat/long calcs

    # Loop through locations (begin as list of 1 NGCC plant location)
    for id, loc in locations.items():
        on_land = loc['on_land'][0]
        survey_rad = land_rad if on_land else osw_rad
        lat = loc['lat'][0]
        lon = loc['lon'][0]
        
        # Add inner circle of 6 surrounding locations @ half of survey radius
        in_circle_lat_delta = [3**.5/4, 0, -(3**.5/4), -(3**.5/4), 0,  3**.5/4]
        for i, lat_delta in enumerate(in_circle_lat_delta):
            lat_delta = survey_rad/earth_rad*lat_delta
            lon_delta = ((survey_rad**2/4/earth_rad**2-lat_delta**2)/\
                            (np.cos(lat/180*np.pi)**2))**.5*180/np.pi
            lat_delta *= 180/np.pi
            if i<3: lon_delta = -lon_delta 
            loc['lat'].append(lat+lat_delta)
            loc['lon'].append(lon+lon_delta)
            if (id == 'WI01') and (i in [3,4]):
                loc['on_land'].append(False)
            elif (id == 'MI02') and (i in [0,1,2]):
                loc['on_land'].append(False)
            elif (id == 'NY01') and (i in [0,1,5]):
                loc['on_land'].append(False)
            elif (id == 'OH03') and (i in [4,5]):
                loc['on_land'].append(False)
            else:
                loc['on_land'].append(globe.is_land(lat+lat_delta,lon+lon_delta))

        # Add outer circle of 12 surrounding location @ full survey radius
        out_circle_lat_delta = [ 1,   3**.5/2,   .5, 0, -.5, -(3**.5/2),
                                -1, -(3**.5/2), -.5, 0,  .5,   3**.5/2]
        for i, lat_delta in enumerate(out_circle_lat_delta):
            lat_delta = survey_rad/earth_rad*lat_delta
            lon_delta = ((survey_rad**2/earth_rad**2-lat_delta**2)/\
                            (np.cos(lat/180*np.pi)**2))**.5*180/np.pi
            lat_delta *= 180/np.pi
            if i<6: lon_delta = -lon_delta 
            loc['lat'].append(lat+lat_delta)
            loc['lon'].append(lon+lon_delta)
            if (id == 'WI01') and (i in [7,8,9,10,11]):
                loc['on_land'].append(False)
            elif (id == 'MI02') and (i in [0,1,2,3,4,5]):
                loc['on_land'].append(False)
            elif (id == 'NY01') and (i in [0,1,2,3,4]):
                loc['on_land'].append(False)
            elif (id == 'OH03') and (i in [9,10,11]):
                loc['on_land'].append(False)
            else:
                loc['on_land'].append(globe.is_land(lat+lat_delta,lon+lon_delta))

    #endregion

    # # Plot survey locations to check
    # #region

    # # Stolen from gis.stackexchange.com/questions/156035
    # def merc_x(lon):
    #   r_major=6378137.000
    #   return r_major*(lon*np.pi/180)
    # def merc_y(lat):
    #   if lat>89.5:lat=89.5
    #   if lat<-89.5:lat=-89.5
    #   r_major=6378137.000
    #   r_minor=6356752.3142
    #   temp=r_minor/r_major
    #   eccent=(1-temp**2)**.5
    #   phi=(lat*np.pi/180)
    #   sinphi=np.sin(phi)
    #   con=eccent*sinphi
    #   com=eccent/2
    #   con=((1.0-con)/(1.0+con))**com
    #   ts=np.tan((np.pi/2-phi)/2)/con
    #   y=0-r_major*np.log(ts)
    #   return y

    # # Set up background image
    # plt.clf
    # bg_img = 'bkg_small.png'
    # img = plt.imread(resource_dir/bg_img)
    # ax = plt.gca()
    # min_x = merc_x(-100)
    # max_x = merc_x(-67)
    # min_y = merc_y(25)
    # max_y = merc_y(46)
    # ax.imshow(img, extent=[min_x, max_x, min_y, max_y])
    # # Plot survey locations
    # for id, loc in locations.items():
    #     for n in range(len(loc['lat'])):
    #         lat = loc['lat'][n]
    #         lon = loc['lon'][n]
    #         color = [1*loc['on_land'][n],0,0]
    #         x = merc_x(lon)
    #         y = merc_y(lat)
    #         ax.plot(x,y,'.',color=color)
    # plt.xlim([min_x,max_x])
    # plt.ylim([min_y,max_y])
    # plt.show()

    # #endregion

    # Add universal financial params to each tech

    for key in finance.keys():
        finance[key]['basis_year'] = sim_basis_year
        finance[key]['plant_lifespan'] = plant_lifespan
        finance[key]['discount_rate'] = discount_rate
        finance[key]['TASC_multiplier'] = TASC_multiplier


    ## Write imported dicts to json dumpfiles

    current_dir = Path(__file__).parent.absolute()
    resource_dir = current_dir/'..'/'resource_files'/'methanol_RCC'/'HOPP_results'/cambium_scenario
    with open(Path(resource_dir/'engin.json'),'w') as file:
        json.dump(engin, file)
    with open(Path(resource_dir/'finance.json'),'w') as file:
        json.dump(finance, file)
    with open(Path(resource_dir/'scenario.json'),'w') as file:
        json.dump(scenario_info, file)
    with open(Path(resource_dir/'locations.json'),'w') as file:
        out_locations = copy.deepcopy(locations)
        for ID, loc in out_locations.items():
            for i, value in enumerate(loc['on_land']):
                value = str(value).lower()
                loc['on_land'][i] = value
        json.dump(out_locations, file)


    # # RUN HOPP HERE - multi_location_RCC.py

    # # %% Check imports and conversions

    # # Load imported dicts from json dumpfiles
    # with open(Path(resource_dir/'engin.json'),'r') as file:
    #     engin = json.load(file)
    # with open(Path(resource_dir/'finance.json'),'r') as file:
    #     finance = json.load(file)
    # with open(Path(resource_dir/'scenario.json'),'r') as file:
    #     scenario_info = json.load(file)
    #     plant_scenarios = scenario_info['plant_scenarios']
    #     cambium_scenarios = scenario_info['cambium_scenarios']
    #     cambium_scenario = scenario_info['cambium_scenario']
    #     atb_scenarios = scenario_info['atb_scenarios']
    #     H2A_scenarios = scenario_info['H2A_scenarios']
    #     MeOH_scenarios = scenario_info['MeOH_scenarios']
    # with open(Path(resource_dir/'locations.json'),'r') as file:
    #     locations = json.load(file)



    # # Print prices per unit for ALL scenarios
    # prices_to_check = ['OCC_$_kw','FOM_$_kwyr','VOM_$_mwh']
    # plants_to_check = plant_scenarios.keys()
    # for plant in plants_to_check:
    #     if plant[0] == 'H':
    #         scenario_list = H2A_scenarios
    #     elif plant[0] == 'M':
    #         scenario_list = MeOH_scenarios
    #     else:
    #         scenario_list = atb_scenarios
    #     for scenario in scenario_list:
    #         for price in prices_to_check:
    #             try:
    #                 price_value = finance[plant][price]
    #                 if type(price_value) is dict:
    #                     price_value = price_value[scenario]
    #                     if type(price_value) is list:
    #                         price_value = price_value[0]
    #                 print('{} plant {} ({} scenario): ${:,.2f}'.format(plant,
    #                                                                 price,
    #                                                                 scenario,
    #                                                                 price_value))
    #             except:
    #                 print('{} plant did not import {} for {} scenario'.format(plant,
    #                                                                         price,
    #                                                                         scenario))

    # # Print absolute prices and parameters for specific scenarios
    # prices_to_check = ['OCC_$','FOM_$_yr','VOM_$_mwh','VOM_other_$_mwh','VOM_elec_$_mwh','VOM_H2_$_mwh','VOM_CO2_$_mwh','VOM_NG_$_mwh','VOM_H2O_$_mwh','VOM_cat_$_mwh']
    # params_to_check = ['output_kw','elec_out_mwh_yr','H2_kg_yr','CO2_kg_yr','MeOH_kg_yr','NG_kg_yr_in'] 
    # for plant in plants_to_check:
    #     scenario = plant_scenarios[plant]
    #     for price in prices_to_check:
    #         try:
    #             price_value = finance[plant][price]
    #             if type(price_value) is dict:
    #                 price_value = price_value[scenario]
    #                 if type(price_value) is list:
    #                     price_value = price_value[0]
    #             print('{} plant {} ({} scenario): ${:,.2f}'.format(plant,
    #                                                             price,
    #                                                             scenario,
    #                                                             price_value))
    #         except:
    #             print('{} plant did not import {} for {} scenario'.format(plant,
    #                                                                     price,
    #                                                                     scenario))
    #     for param in params_to_check:
    #         # if 'prod' in param:
    #         #     if 'CC' in plant:
    #         #         product = 'CO2'
    #         #     elif plant[0] == 'H':
    #         #         product = 'H2'
    #         #     elif plant[0] == 'M':
    #         #         product = 'MeOH'
    #         #     else:
    #         #         product = 'elec'
    #         #     if product == 'elec':
    #         #         param = 'elec_out_mwh_yr'
    #         #     else:
    #         #         param = param = product+'_kg_yr'
    #         try:
    #             param_value = engin[plant][param]
    #             if type(param_value) is dict:
    #                 param_value = param_value[scenario]
    #                 if type(param_value) is list:
    #                     param_value = param_value[0]
    #             print('{} plant {} ({} scenario): {:,.2f}'.format(plant,
    #                                                             param,
    #                                                             scenario,
    #                                                             param_value))
    #         except:
    #             print('{} plant did not import {} for {} scenario'.format(plant,
    #                                                                     param,
    #                                                                     scenario))

    ##
