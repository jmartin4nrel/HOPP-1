from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from pathlib import Path
import numpy as np
import json
from global_land_mask import globe

# Physical constants

C_MW = 12.011 # g/mol C
H_MW = 1.008 # g/mol H
N_MW = 14.007 # g/mol N
O_MW = 15.999 # g/mol O

CH4_MW = C_MW+H_MW*4
C2H6_MW = C_MW*2+H_MW*6
C3H8_MW = C_MW*3+H_MW*8
C4H10_MW = C_MW*4+H_MW*10
CO2_MW = C_MW+O_MW*2
N2_MW = N_MW*2

# From NETL-PUB-22638, p. 510
ng_MW = .931*CH4_MW+\
        .032*C2H6_MW+\
        .007*C3H8_MW+\
        .004*C4H10_MW+\
        .01*CO2_MW+\
        .016*N2_MW
mol_C_mol_ng = .931+.032*2+.007*3+.004*4+.01
kg_co2_kg_ng = mol_C_mol_ng*CO2_MW/ng_MW

# From NETL-PUB-22638, p. 514
no_ccs_ng_in_kg_hr = 67369*ng_MW/C_MW
no_ccs_co2_out_kg_hr = 67890*CO2_MW/C_MW
no_ccs_kg_co2_kg_ng = no_ccs_co2_out_kg_hr/no_ccs_ng_in_kg_hr

# From NETL-PUB-22638, p. 514
ccs_ng_in_kg_hr = 67369*ng_MW/C_MW
ccs_co2_out_kg_hr = 61090*CO2_MW/C_MW
ccs_kg_co2_kg_ng = ccs_co2_out_kg_hr/ccs_ng_in_kg_hr

kJ_btu = 1.05506 # kJ/BTU
NG_LHV_MJ_kg = 47.21 # Natural gas net calorific value, MJ/kg,
# Calculated from composition above and https://www.unitrove.com/engineering/tools/gas/natural-gas-calorific-value

ccs_eff = 0.9 # carbon capture efficiency - 90%
ngcc_cap_fac = 0.85 # NGCC plant capacity factor - 85%

def load_atb_dicts():

    '''
    Creates .json files containing renewable resource parameters determined by the NREL Annual Technology Baseline (ATB)
        -engin.json: engineering constants
        -finance.json: financial constants
    '''
    
    # Set the filename, sheets, scenarios, and columns for this instance of the ATB
    filename = "2022 v3 Annual Technology Baseline Workbook Corrected 1-24-2023.xlsx"
    atb_scenarios = ['Advanced','Moderate','Conservative']
    sheets = {'NGCC':'Natural Gas_FE',
            'CCS': 'Natural Gas_FE',
            'PV':  'Solar - Utility PV',
            'LBW': 'Land-Based Wind',
            'OSW': 'Offshore Wind'}
    base_cols =    {'NGCC':'M',
                    'CCS': 'M',
                    'PV':  'M',
                    'LBW': 'M',
                    'OSW': 'M'}
    year_rows =    {'NGCC':68,
                    'CCS': 68,
                    'PV':  74,
                    'LBW': 75,
                    'OSW': 76}

    # Set the rows that engineering constants are found on each sheet
    engin ={'NGCC':{'heat_rate_btu_wh':{'Advanced':72,
                                        'Moderate':73,
                                        'Conservative':74}},
            'CCS': {'heat_rate_btu_wh':{'Advanced':75,
                                        'Moderate':76,
                                        'Conservative':77}},
            'PV':  {'bifaciality':     {'Advanced':75,
                                        'Moderate':76,
                                        'Conservative':77},
                    'albedo':          {'Advanced':78,
                                        'Moderate':79,
                                        'Conservative':80},
                    'losses':          {'Advanced':81,
                                        'Moderate':82,
                                        'Conservative':83},
                    'dc_degradation':  {'Advanced':84,
                                        'Moderate':85,
                                        'Conservative':86},
                    'inv_eff':         {'Advanced':87,
                                        'Moderate':88,
                                        'Conservative':89},
                    'capacity_factor': {'Advanced':90,
                                        'Moderate':91,
                                        'Conservative':92}},
            'LBW': {'hub_height':      {'Advanced':76,
                                        'Moderate':77,
                                        'Conservative':78},
                    'rotor_diameter':  {'Advanced':79,
                                        'Moderate':80,
                                        'Conservative':81},
                    'turbine_rating_kw':{'Advanced':82,
                                        'Moderate':83,
                                        'Conservative':84},
                    'capacity_factor': {'Advanced':85,
                                        'Moderate':86,
                                        'Conservative':87}},
            'OSW': {'hub_height':      {'Advanced':77,
                                        'Moderate':78,
                                        'Conservative':79},
                    'rotor_diameter':  {'Advanced':80,
                                        'Moderate':81,
                                        'Conservative':82},
                    'turbine_rating_kw':{'Advanced':83,
                                        'Moderate':84,
                                        'Conservative':85},
                    'capacity_factor': {'Advanced':86,
                                        'Moderate':87,
                                        'Conservative':88}}}

    # Set the rows that financial constants are found on each sheet
    # toc: Total Overnight Cost ($/kW), foc: Fixed O&M Cost ($/kW-yr), voc: Variable O&M Cost ($/kWh)
    # NOTE: voc was changed from original units of $/MWh in the raw ATB download to match simple financial model
    finance =  {'NGCC':{'toc_kw':    {'Advanced':105,
                                        'Moderate':106,
                                        'Conservative':107},
                        'foc_kw_yr':  {'Advanced':116,
                                        'Moderate':117,
                                        'Conservative':118},
                        'voc_kwh':   {'Advanced':127,
                                        'Moderate':128,
                                        'Conservative':129}},
                'CCS': {'toc_kw':    {'Advanced':108,
                                        'Moderate':109,
                                        'Conservative':110},
                        'foc_kw_yr':  {'Advanced':119,
                                        'Moderate':120,
                                        'Conservative':121},
                        'voc_kwh':   {'Advanced':130,
                                        'Moderate':131,
                                        'Conservative':132}},
                'PV':  {'toc_kw':    {'Advanced':203,
                                        'Moderate':204,
                                        'Conservative':205},
                        'foc_kw_yr':  {'Advanced':235,
                                        'Moderate':236,
                                        'Conservative':237},
                        'voc_kwh':   {'Advanced':267,
                                        'Moderate':268,
                                        'Conservative':269}},
                'LBW': {'toc_kw':    {'Advanced':204,
                                        'Moderate':205,
                                        'Conservative':206},
                        'foc_kw_yr':  {'Advanced':236,
                                        'Moderate':237,
                                        'Conservative':238},
                        'voc_kwh':   {'Advanced':268,
                                        'Moderate':269,
                                        'Conservative':270}},
                'OSW': {'toc_kw':    {'Advanced':253,
                                        'Moderate':254,
                                        'Conservative':255},
                        'foc_kw_yr':  {'Advanced':297,
                                        'Moderate':298,
                                        'Conservative':299},
                        'VOM_$_mwh':   {'Advanced':341,
                                        'Moderate':342,
                                        'Conservative':343}}}

    # Import from spreadsheet
    atb_basis_year = 2020
    atb_inflation_rate = 0.025
    resource_dir = Path(__file__).parent.absolute()/'inputs'
    workbook = load_workbook(resource_dir/filename, read_only=True, data_only=True)
    atb_tech = ['NGCC','CCS','PV','LBW','OSW']

    sim_basis_year = 2020
    sim_start_year = 2020
    sim_end_year = 2050
    sim_increment = 5
    sim_years = np.arange(sim_start_year,sim_end_year+sim_increment,sim_increment)

    for tech in atb_tech:
        sheet_name = sheets[tech]
        worksheet = workbook[sheet_name]
        # Figure out column range for years
        base_col = base_cols[tech]
        base_col_idx = column_index_from_string(base_col)
        year_row = str(year_rows[tech])
        base_year = worksheet[base_col+year_row].value
        start_col_idx = base_col_idx+sim_start_year-base_year
        end_col_idx = base_col_idx+sim_end_year-base_year
        start_col = get_column_letter(start_col_idx)
        end_col = get_column_letter(end_col_idx)
        # Replace row numbers in nested dicts with engin/finance params
        for param_dict in [engin, finance]:
            for param in param_dict[tech].values():
                for scenario, row in param.items():
                    cells = worksheet[start_col+str(row)+':'+end_col+str(row)][0]
                    values = []
                    for i, cell in enumerate(cells):
                        if i+base_year in sim_years:
                            value = cell.value
                            if param_dict is finance:
                                value = value*(1+atb_inflation_rate)**(sim_basis_year-atb_basis_year)
                            values.append(value)
                    param[scenario] = values

    pv_constants = {'array_type':2,
                    'azimuth':180,
                    'dc_ac_ratio':1.28} # ILR NOT OPTIMIZED - can HOPP do this?
    for constant in pv_constants:
        engin['PV'][constant] = {}
        for scenario in atb_scenarios:
            engin['PV'][constant][scenario] = [pv_constants[constant] for _ in sim_years]

    wind_constants =   {'wind_turbine_max_cp':0.45,
                        'avail_bop_loss':0,
                        'avail_grid_loss':0,
                        'avail_turb_loss':0,
                        'elec_eff_loss':0,
                        'elec_parasitic_loss':0,
                        'env_degrad_loss':0,
                        'env_env_loss':0,
                        'env_icing_loss':0,
                        'ops_env_loss':0,
                        'ops_grid_loss':0,
                        'ops_load_loss':0,
                        'turb_generic_loss':0,
                        'turb_hysteresis_loss':0,
                        'turb_perf_loss':0,
                        'turb_specific_loss':0,
                        'wake_ext_loss':0}
    for constant in wind_constants:
        engin['LBW'][constant] = {}
        engin['OSW'][constant] = {}
        for scenario in atb_scenarios:
            engin['LBW'][constant][scenario] = [wind_constants[constant] for _ in sim_years]
            engin['OSW'][constant][scenario] = [wind_constants[constant] for _ in sim_years]

    out_path = resource_dir/'atb_engin.json'
    with open(out_path, 'w') as outfile:
        json.dump(engin, outfile)
    out_path = resource_dir/'atb_finance.json'
    with open(out_path, 'w') as outfile:
        json.dump(finance, outfile)


def set_atb_year(hi, scenario, year):

    year_idx = int((year-2020)/5)

    resource_dir = Path(__file__).parent.absolute()/'inputs'
    in_path = resource_dir/'atb_engin.json'
    with open(in_path, 'r') as infile:
        engin = json.load(infile)
    in_path = resource_dir/'atb_finance.json'
    with open(in_path, 'r') as infile:
        finance = json.load(infile)
    

    # Determine if location is on land
    on_land = globe.is_land(hi.system.site.data['lat'], hi.system.site.data['lon'])

    techs = ['wind','pv','co2']
    tech_list = list(hi.system.tech_config._get_model_dict().keys())
    for tech in techs:
        if tech in tech_list:
            tech_finance = {}
            tech_engin = {}
            if tech == 'wind':
                if not on_land:
                    for key in list(finance['OSW'].keys()):
                        tech_finance[key] = finance['OSW'][key][scenario]
                    for key in list(engin['OSW'].keys()):
                        tech_engin[key] = engin['OSW'][key][scenario]
                else:
                    for key in list(finance['LBW'].keys()):
                        tech_finance[key] = finance['LBW'][key][scenario]
                    for key in list(engin['LBW'].keys()):
                        tech_engin[key] = engin['LBW'][key][scenario]
                for key, value in tech_finance.items():
                    setattr(hi.system.wind._financial_model,key,value[year_idx])
                hi.system.wind.turb_rating = tech_engin['turbine_rating_kw'][year_idx]
                hi.system.wind.rotor_diameter = tech_engin['rotor_diameter'][year_idx]
                hi.system.wind._system_model.wind_turbine_max_cp = tech_engin['wind_turbine_max_cp'][year_idx]
                hi.system.wind._system_model.wind_turbine_hub_ht = tech_engin['hub_height'][year_idx]
                hi.system.wind._system_model.wind_turbine_rotor_diameter = tech_engin['rotor_diameter'][year_idx]
                for key, value in tech_engin.items():
                    if key not in ['wind_turbine_max_cp','hub_height','rotor_diameter','turbine_rating_kw','capacity_factor']:
                        setattr(hi.system.wind._system_model.Losses,key,value[year_idx])
            elif tech == 'pv':
                for key in list(finance['PV'].keys()):
                    tech_finance[key] = finance['PV'][key][scenario]
                for key in list(engin['PV'].keys()):
                    tech_engin[key] = engin['PV'][key][scenario]
                for key, value in tech_finance.items():
                    setattr(hi.system.pv._financial_model,key,value[year_idx])
                hi.system.pv._system_model.Lifetime.dc_degradation = tuple([tech_engin['dc_degradation'][year_idx]])
                hi.system.pv._system_model.SolarResource.albedo = tuple(np.ones(12)*tech_engin['albedo'][year_idx])
                for key, value in tech_engin.items():
                    if key not in ['albedo','dc_degradation','capacity_factor']:
                        setattr(hi.system.pv._system_model.SystemDesign,key,value[year_idx])
            elif tech == 'co2':
                no_ccs_finance = {}
                no_ccs_engin = {}
                ccs_finance = {}
                ccs_engin = {}
                for key in list(finance['NGCC'].keys()):
                    no_ccs_finance[key] = finance['NGCC'][key][scenario]
                for key in list(engin['NGCC'].keys()):
                    no_ccs_engin[key] = engin['NGCC'][key][scenario]
                for key in list(finance['CCS'].keys()):
                    ccs_finance[key] = finance['CCS'][key][scenario]
                for key in list(engin['CCS'].keys()):
                    ccs_engin[key] = engin['CCS'][key][scenario]

                no_ccs_co2_kg_kwh = [i*kJ_btu/NG_LHV_MJ_kg*no_ccs_kg_co2_kg_ng for i in no_ccs_engin['heat_rate_btu_wh']]#*0+1/2.9225
                ccs_co2_kg_kwh = [i*kJ_btu/NG_LHV_MJ_kg*ccs_kg_co2_kg_ng for i in ccs_engin['heat_rate_btu_wh']]#*0+1/2.886

                heat_rate_ratio = []
                for i in range(len(ccs_co2_kg_kwh)):
                    heat_rate_ratio.append(ccs_engin['heat_rate_btu_wh'][i]/no_ccs_engin['heat_rate_btu_wh'][i])

                setattr(hi.system.co2._system_model,'no_ccs_co2_kg_kwh',no_ccs_co2_kg_kwh[year_idx])
                setattr(hi.system.co2._system_model,'ccs_co2_kg_kwh',ccs_co2_kg_kwh[year_idx])
                setattr(hi.system.co2._system_model,'no_ccs_kg_co2_kg_ng',no_ccs_kg_co2_kg_ng)
                setattr(hi.system.co2._system_model,'ccs_kg_co2_kg_ng',ccs_kg_co2_kg_ng)
                
                for key, ccs_value in ccs_finance.items():
                    no_ccs_value = no_ccs_finance[key]
                    value = []
                    if key == 'toc_kw':
                        for i in range(len(ccs_value)):
                            value.append((ccs_value[i]*heat_rate_ratio[i]-no_ccs_value[i])/ccs_co2_kg_kwh[i]*3600)#
                        setattr(hi.system.co2._financial_model,'toc_kg_s',value[year_idx])
                    elif key == 'foc_kw_yr':
                        for i in range(len(ccs_value)):
                            value.append((ccs_value[i]*heat_rate_ratio[i]-no_ccs_value[i])/ccs_co2_kg_kwh[i]*3600)#
                        setattr(hi.system.co2._financial_model,'foc_kg_s_yr',value[year_idx])
                    elif key == 'voc_kwh':
                        for i in range(len(ccs_value)):
                            value.append((ccs_value[i]-no_ccs_value[i])/ccs_co2_kg_kwh[i])#
                        setattr(hi.system.co2._financial_model,'voc_kg',value[year_idx])

    return hi


if __name__ == '__main__':

    load_atb_dicts()